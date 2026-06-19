"""
InfecSure — Report Generation Service
=======================================
Generates PDF and Excel executive reports.

Heavy libraries (reportlab, openpyxl) are imported lazily inside each
generator function so they only load when a report is actually requested,
keeping idle RAM low on Render's free tier.
"""

from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Reports are saved to a temp directory (or Firebase Storage in production)
REPORTS_DIR = Path("reports_output")
REPORTS_DIR.mkdir(exist_ok=True)


# ─── PDF Generation ───────────────────────────────────────────────────────────

def _get_pdf_styles():
    """Build ReportLab paragraph styles. Called only from within PDF functions."""
    from reportlab.lib import colors  # noqa: PLC0415
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # noqa: PLC0415
    from reportlab.lib.enums import TA_CENTER  # noqa: PLC0415

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="InfecSureTitle",
        fontSize=18, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1a1a2e"),
        spaceAfter=6, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="InfecSureSubtitle",
        fontSize=11, fontName="Helvetica",
        textColor=colors.HexColor("#444"),
        spaceAfter=12, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="SectionHeader",
        fontSize=13, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1a1a2e"),
        spaceBefore=12, spaceAfter=6,
    ))
    return styles


def generate_executive_pdf(
    wards: list[dict],
    alerts: list[dict],
    audit_summary: list[dict],
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    generated_by: str = "ICNO",
) -> bytes:
    """
    Generate an Executive Summary PDF report.
    Returns raw PDF bytes.
    """
    # ── Lazy imports — only loaded when this function is called ────────────
    try:
        from reportlab.lib import colors  # noqa: PLC0415
        from reportlab.lib.pagesizes import A4  # noqa: PLC0415
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError:
        raise RuntimeError("ReportLab is not installed. Run: pip install reportlab")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = _get_pdf_styles()
    story = []

    # ── Header ────────────────────────────────────────────────────────────
    story.append(Paragraph("InfecSure", styles["InfecSureTitle"]))
    story.append(Paragraph("AI-Powered Infection Monitoring System", styles["InfecSureSubtitle"]))
    story.append(Paragraph("Executive Summary Report", styles["InfecSureSubtitle"]))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1a1a2e")))
    story.append(Spacer(1, 0.3*cm))

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story.append(Paragraph(f"Generated: {now_str} | By: {generated_by}", styles["Normal"]))
    if date_from and date_to:
        story.append(Paragraph(
            f"Period: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}",
            styles["Normal"],
        ))
    story.append(Spacer(1, 0.5*cm))

    # ── Ward Risk Summary ─────────────────────────────────────────────────
    story.append(Paragraph("Ward Risk Overview", styles["SectionHeader"]))
    ward_data = [["Ward Name", "Type", "Risk Level", "Risk Score", "Compliance %"]]
    for w in wards:
        risk = w.get("risk_level", "low").upper()
        ward_data.append([
            w.get("name", "—"),
            w.get("ward_type", "—").title(),
            risk,
            f"{w.get('risk_score', 0.0):.1f}",
            f"{w.get('compliance_score', 100.0):.1f}%",
        ])

    risk_table = Table(ward_data, colWidths=[4.5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    risk_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(risk_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Alerts Summary ────────────────────────────────────────────────────
    story.append(Paragraph("Validated Alerts", styles["SectionHeader"]))
    if alerts:
        alert_data = [["Title", "Severity", "Ward", "Status", "Date"]]
        for a in alerts[:20]:  # Cap at 20
            alert_data.append([
                a.get("title", "—")[:40],
                a.get("severity", "—").upper(),
                a.get("ward_id", "—")[:20],
                a.get("status", "—").upper(),
                str(a.get("created_at", "—"))[:10],
            ])
        alert_table = Table(alert_data, colWidths=[5*cm, 2.5*cm, 3*cm, 2.5*cm, 3.5*cm])
        alert_table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#16213e")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(alert_table)
    else:
        story.append(Paragraph("No validated alerts in this period.", styles["Normal"]))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Paragraph(
        "InfecSure v1.0 — Divisional Hospital, Thalangama, Colombo, Sri Lanka",
        styles["Normal"],
    ))

    doc.build(story)
    return buffer.getvalue()


# ─── Excel Generation ─────────────────────────────────────────────────────────

def generate_executive_excel(
    wards: list[dict],
    alerts: list[dict],
    lab_results: list[dict],
) -> bytes:
    """
    Generate an Executive Summary Excel workbook.
    Returns raw .xlsx bytes.
    """
    # ── Lazy imports ───────────────────────────────────────────────────────
    try:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: PLC0415
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def style_data_row(ws, row: int, cols: int):
        fill = PatternFill("solid", fgColor="F5F5F5" if row % 2 == 0 else "FFFFFF")
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    # ── Sheet 1: Ward Risk ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ward Risk Summary"
    headers1 = ["Ward Name", "Type", "Risk Level", "Risk Score", "Compliance %", "Last Audit"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header_row(ws1, 1, len(headers1))

    for r, w in enumerate(wards, 2):
        ws1.cell(row=r, column=1, value=w.get("name"))
        ws1.cell(row=r, column=2, value=w.get("ward_type", "").title())
        ws1.cell(row=r, column=3, value=w.get("risk_level", "low").upper())
        ws1.cell(row=r, column=4, value=round(w.get("risk_score", 0.0), 1))
        ws1.cell(row=r, column=5, value=round(w.get("compliance_score", 100.0), 1))
        ws1.cell(row=r, column=6, value=str(w.get("last_audit_at", ""))[:10])
        style_data_row(ws1, r, len(headers1))

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 18

    # ── Sheet 2: Alerts ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Alerts")
    headers2 = ["Title", "Type", "Severity", "Ward", "Status", "ICNO Notes", "Date"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers2))

    for r, a in enumerate(alerts, 2):
        ws2.cell(row=r, column=1, value=a.get("title"))
        ws2.cell(row=r, column=2, value=a.get("alert_type"))
        ws2.cell(row=r, column=3, value=a.get("severity", "").upper())
        ws2.cell(row=r, column=4, value=a.get("ward_id"))
        ws2.cell(row=r, column=5, value=a.get("status", "").upper())
        ws2.cell(row=r, column=6, value=a.get("icno_notes", ""))
        ws2.cell(row=r, column=7, value=str(a.get("created_at", ""))[:10])
        style_data_row(ws2, r, len(headers2))

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 20

    # ── Sheet 3: Lab Results ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Lab Results")
    headers3 = ["Ward", "Pathogen", "Specimen", "Date", "Anomaly", "Z-Score"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header_row(ws3, 1, len(headers3))

    for r, lr in enumerate(lab_results, 2):
        anomaly = lr.get("anomaly") or {}
        ws3.cell(row=r, column=1, value=lr.get("ward_id"))
        ws3.cell(row=r, column=2, value=lr.get("pathogen_name"))
        ws3.cell(row=r, column=3, value=lr.get("specimen_type", "").title())
        ws3.cell(row=r, column=4, value=str(lr.get("result_date", ""))[:10])
        ws3.cell(row=r, column=5, value="YES" if anomaly.get("is_anomaly") else "NO")
        ws3.cell(row=r, column=6, value=anomaly.get("z_score", 0.0))
        style_data_row(ws3, r, len(headers3))

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = 18

    # Return bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Dengue PDF Report ────────────────────────────────────────────────────────

def generate_dengue_pdf(alert: dict, lab_results: list[dict], generated_by: str, audits: Optional[list[dict]] = None) -> bytes:
    """Generate a formatted Dengue Alert Report PDF for the Supervising Doctor."""
    # ── Lazy imports ───────────────────────────────────────────────────────
    try:
        from reportlab.lib import colors  # noqa: PLC0415
        from reportlab.lib.pagesizes import A4  # noqa: PLC0415
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError:
        raise RuntimeError("ReportLab is not installed.")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = _get_pdf_styles()
    story = []

    story.append(Paragraph("InfecSure — Dengue Alert Report", styles["InfecSureTitle"]))
    story.append(Paragraph(
        f"ICNO Validated | {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        styles["InfecSureSubtitle"],
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#dc3545")))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(f"Alert Title: {alert.get('title', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"Severity: {alert.get('severity', 'N/A').upper()}", styles["Normal"]))
    story.append(Paragraph(f"Description: {alert.get('description', 'N/A')}", styles["Normal"]))
    if alert.get("icno_notes"):
        story.append(Paragraph(f"ICNO Notes: {alert['icno_notes']}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Related ICNO Ward Audits", styles["SectionHeader"]))
    if audits:
        audit_data = [["Ward", "Overall", "Hand Hygiene", "PPE", "Waste", "Environment", "Date"]]
        for audit in audits[:20]:
            audit_data.append([
                audit.get("ward_id", "N/A"),
                f"{float(audit.get('overall_compliance_score', 0) or 0):.1f}%",
                f"{float(audit.get('hand_hygiene_score', 0) or 0):.1f}%",
                f"{float(audit.get('ppe_score', 0) or 0):.1f}%",
                f"{float(audit.get('waste_segregation_score', 0) or 0):.1f}%",
                f"{float(audit.get('environmental_score', 0) or 0):.1f}%",
                str(audit.get("created_at", audit.get("audit_date", "")))[:10],
            ])
        audit_table = Table(audit_data, colWidths=[3*cm, 2.2*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm])
        audit_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(audit_table)
    else:
        story.append(Paragraph("No related ICNO ward audits found for this period.", styles["Normal"]))

    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Related Lab Results", styles["SectionHeader"]))
    if lab_results:
        lr_data = [["Pathogen", "Specimen", "Date", "Anomaly", "Z-Score"]]
        for lr in lab_results[:30]:
            anomaly = lr.get("anomaly") or {}
            lr_data.append([
                lr.get("pathogen_name", "—"),
                lr.get("specimen_type", "—").title(),
                str(lr.get("result_date", "—"))[:10],
                "YES" if anomaly.get("is_anomaly") else "NO",
                f"{anomaly.get('z_score', 0.0):.2f}",
            ])
        t = Table(lr_data, colWidths=[4*cm, 3*cm, 3*cm, 2.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#dc3545")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No lab results linked.", styles["Normal"]))

    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Please acknowledge this report in InfecSure and issue management instructions.", styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()
